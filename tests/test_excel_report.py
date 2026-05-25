from pathlib import Path

from openpyxl import load_workbook

from models.schemas import Product
from utils.excel_report import build_excel_row, write_products_excel_report


def test_build_excel_row_maps_product_into_template_like_columns() -> None:
    product = Product(
        id="4225897",
        name="Вино Free Feather Chardonnay безалкогольное полусладкое белое 750мл",
        brand="Free Feather",
        price=699.99,
        image_url="https://img.example/4225897.webp",
        product_link="https://5ka.ru/product/vino-free-feather--4225897/",
        category="Безалкогольное вино",
        subcategory="Тихое",
        in_stock=True,
        raw_data={"alcohol_type": "Безалкогольное"},
    )

    row = build_excel_row(product, shop="pyaterochka", row_number=1)

    assert row["№№"] == 1
    assert row["Классификация по остаточному сахару"] == "Полусладкое"
    assert row["Классификация по цвету, годам"] == "Белое"
    assert row["Наименование"] == product.name
    assert row["Винодельня"] == "Free Feather"
    assert row["Объем, л"] == 0.75
    assert row["Рекомендуемая стоимость на полке в пределах агрегированной цены (руб.)"] == 699.99
    assert row["Агрегированная стоимость (руб.)"] == 699.99
    assert row["Наш магазин"] == "pyaterochka"
    assert row["Наш тип вина"] == "Тихое"
    assert row["Наш алкогольный тип"] == "Безалкогольное"
    assert row["Наш производитель/поставщик"] == "Free Feather"
    assert row["Ссылка на вино"] == "https://5ka.ru/product/vino-free-feather--4225897/"
    assert row["Наша цена"] == 699.99


def test_write_products_excel_report_creates_workbook_with_summary_and_category_sheets(
    tmp_path: Path,
) -> None:
    products = [
        Product(
            id="4225897",
            name="Вино Free Feather Chardonnay безалкогольное полусладкое белое 750мл",
            brand="Free Feather",
            price=699.99,
            image_url="https://img.example/4225897.webp",
            product_link="https://5ka.ru/product/vino-free-feather--4225897/",
            category="Безалкогольное вино",
            subcategory="Тихое",
            in_stock=True,
            raw_data={"alcohol_type": "Безалкогольное"},
        ),
        Product(
            id="4015936",
            name="Горбуша дальневосточная Polar филе порционное замороженное 400г",
            price=519.99,
            image_url="https://img.example/4015936.webp",
            product_link="https://5ka.ru/product/gorbusha-dalnevostochnaya-polar--4015936/",
            category="Рыба",
            in_stock=True,
        ),
    ]

    report_path = write_products_excel_report(
        products,
        shop="pyaterochka",
        output_dir=tmp_path,
        exported_at="2026-05-20T09:30:00",
    )

    assert isinstance(report_path, Path)
    assert report_path.exists()
    assert report_path.suffix == ".xlsx"

    workbook = load_workbook(report_path, read_only=True)
    assert workbook.sheetnames == ["Сводка", "Безалкогольное вино", "Рыба"]

    summary_sheet = workbook["Сводка"]
    summary_rows = list(summary_sheet.iter_rows(min_row=1, max_row=5, values_only=True))
    assert ("Магазин", "pyaterochka") in summary_rows
    assert ("Товаров", 2) in summary_rows

    wine_sheet = workbook["Безалкогольное вино"]
    wine_headers = next(wine_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    wine_row = next(wine_sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    assert wine_headers[0] == "№№"
    assert "Наш тип вина" in wine_headers
    assert "Наш алкогольный тип" in wine_headers
    assert "Наш производитель/поставщик" in wine_headers
    assert "Ссылка на вино" in wine_headers
    assert wine_row[0] == 1
    assert wine_row[3] == "Вино Free Feather Chardonnay безалкогольное полусладкое белое 750мл"

    fish_sheet = workbook["Рыба"]
    fish_row = next(fish_sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    assert fish_row[0] == 1
    assert fish_row[3] == "Горбуша дальневосточная Polar филе порционное замороженное 400г"
