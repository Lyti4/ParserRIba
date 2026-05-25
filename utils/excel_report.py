"""Excel report generation for exported store products."""

from __future__ import annotations

import re
from collections import OrderedDict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from models.schemas import Product

NETWORK_COLUMNS = (
    "Деликатеска",
    "Глобус",
    "Метро",
    "Перекресток",
    "Ашан",
    "Окей",
    "Лента",
    "Азбука вкуса",
    "Винстайл",
    "Красное белое",
    "Ароматный мир",
    "SimpleWine",
    "Винлаб",
    "КУПЕР",
    "Прочие магазины",
)

BASE_COLUMNS = [
    "№№",
    "Классификация по остаточному сахару",
    "Классификация по цвету, годам",
    "Наименование",
    "Винодельня",
    "Контакт поставщика",
    "Объем, л",
    "Закупочная стоимость товара (руб.)",
    "Рекомендуемая стоимость на полке в пределах агрегированной цены (руб.)",
    "Наценка (%)",
    "Агрегированная стоимость (руб.)",
]
NETWORK_PAIR_COLUMNS = [
    item
    for network in NETWORK_COLUMNS
    for item in (f"{network} цена", f"{network} ссылка")
]
EXTRA_COLUMNS = [
    "Наш магазин",
    "Наша категория",
    "Наш тип вина",
    "Наш алкогольный тип",
    "Наш производитель/поставщик",
    "Наш бренд",
    "Наш товар ID",
    "Наша цена",
    "Наша старая цена",
    "Наша цена за единицу",
    "Ссылка на вино",
    "Наше изображение",
    "В наличии",
    "Экспортировано",
]
REPORT_COLUMNS = BASE_COLUMNS + NETWORK_PAIR_COLUMNS + EXTRA_COLUMNS

SHOP_TO_NETWORK = {
    "metro": "Метро",
    "perekrestok": "Перекресток",
    "auchan": "Ашан",
    "okey": "Окей",
    "lenta": "Лента",
    "azbukavkusa": "Азбука вкуса",
    "krasnoeibeloe": "Красное белое",
}

SUMMARY_ROWS = (
    ("Магазин", "shop"),
    ("Товаров", "products_count"),
    ("Категорий", "categories_count"),
    ("Категории", "categories"),
    ("Экспортировано", "exported_at"),
)


def build_excel_row(
    product: Product,
    *,
    shop: str,
    row_number: int,
    exported_at: str = "",
) -> dict[str, object]:
    """Build one template-like Excel row from a Product."""
    row = {column: "" for column in REPORT_COLUMNS}
    row["№№"] = row_number
    row["Классификация по остаточному сахару"] = _extract_sugar_class(product.name)
    row["Классификация по цвету, годам"] = _extract_color_and_years(product.name)
    row["Наименование"] = product.name
    row["Винодельня"] = str(product.brand or "")
    row["Контакт поставщика"] = _extract_supplier(product)
    row["Объем, л"] = _extract_volume_liters(product)
    row["Рекомендуемая стоимость на полке в пределах агрегированной цены (руб.)"] = float(
        product.price.current
    )
    row["Агрегированная стоимость (руб.)"] = float(product.price.current)

    network_name = SHOP_TO_NETWORK.get(str(shop or "").casefold())
    if network_name:
        row[f"{network_name} цена"] = float(product.price.current)
        row[f"{network_name} ссылка"] = str(product.product_link)

    alcohol_type = _extract_alcohol_type(product)
    row["Наш магазин"] = shop
    row["Наша категория"] = str(product.category or "")
    row["Наш тип вина"] = str(product.subcategory or "")
    row["Наш алкогольный тип"] = alcohol_type
    row["Наш производитель/поставщик"] = _extract_supplier(product)
    row["Наш бренд"] = str(product.brand or "")
    row["Наш товар ID"] = str(product.id or "")
    row["Наша цена"] = float(product.price.current)
    row["Наша старая цена"] = product.price.old if product.price.old is not None else ""
    row["Наша цена за единицу"] = product.price.unit if product.price.unit is not None else ""
    row["Ссылка на вино"] = str(product.product_link)
    row["Наше изображение"] = str(product.image_url or "")
    row["В наличии"] = "Да" if product.in_stock else "Нет"
    row["Экспортировано"] = exported_at
    return row


def write_products_excel_report(
    products: list[Product],
    *,
    shop: str,
    output_dir: Path | str,
    exported_at: str = "",
) -> Path:
    """Write products into an XLSX workbook with summary and category sheets."""
    target_dir = Path(output_dir)
    target_dir.mkdir(parents=True, exist_ok=True)
    report_path = target_dir / f"{shop}_products.xlsx"

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Сводка"

    categories = _group_products_by_category(products)
    _write_summary_sheet(
        summary_sheet,
        shop=shop,
        categories=categories,
        exported_at=exported_at,
    )

    for category_name, category_products in categories.items():
        sheet = workbook.create_sheet(title=_sheet_title(category_name))
        _write_products_sheet(
            sheet,
            products=category_products,
            shop=shop,
            exported_at=exported_at,
        )

    workbook.save(report_path)
    return report_path


def _write_summary_sheet(
    sheet: Worksheet,
    *,
    shop: str,
    categories: OrderedDict[str, list[Product]],
    exported_at: str,
) -> None:
    values = {
        "shop": shop,
        "products_count": sum(len(items) for items in categories.values()),
        "categories_count": len(categories),
        "categories": ", ".join(categories.keys()),
        "exported_at": exported_at,
    }
    for index, (label, key) in enumerate(SUMMARY_ROWS, start=1):
        sheet.cell(row=index, column=1, value=label).font = Font(bold=True)
        sheet.cell(row=index, column=2, value=values[key])
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 80


def _write_products_sheet(
    sheet: Worksheet,
    *,
    products: list[Product],
    shop: str,
    exported_at: str,
) -> None:
    sheet.append(REPORT_COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for index, product in enumerate(products, start=1):
        row = build_excel_row(product, shop=shop, row_number=index, exported_at=exported_at)
        sheet.append([row[column] for column in REPORT_COLUMNS])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    _apply_column_widths(sheet)


def _group_products_by_category(products: list[Product]) -> OrderedDict[str, list[Product]]:
    grouped: OrderedDict[str, list[Product]] = OrderedDict()
    for product in products:
        category_name = str(product.category or "Без категории")
        grouped.setdefault(category_name, []).append(product)
    return grouped


def _sheet_title(name: str) -> str:
    sanitized = re.sub(r"[:\\/?*\[\]]", "_", str(name or "Без категории")).strip()
    if not sanitized:
        sanitized = "Без категории"
    return sanitized[:31]


def _apply_column_widths(sheet: Worksheet) -> None:
    widths = {
        "A": 8,
        "B": 24,
        "C": 24,
        "D": 48,
        "E": 24,
        "F": 24,
        "G": 10,
        "H": 18,
        "I": 18,
        "J": 12,
        "K": 18,
    }
    for column_letter, width in widths.items():
        sheet.column_dimensions[column_letter].width = width


def _extract_supplier(product: Product) -> str:
    raw_data = dict(product.raw_data or {})
    for key in ("supplier", "producer", "manufacturer", "vendor", "brand"):
        value = raw_data.get(key)
        if isinstance(value, str):
            text = value.strip()
            if text:
                return text
    return str(product.brand or "")


def _extract_alcohol_type(product: Product) -> str:
    raw_data = dict(product.raw_data or {})
    value = raw_data.get("alcohol_type")
    if isinstance(value, str):
        return value.strip()
    return ""


def _extract_sugar_class(name: str) -> str:
    lowered = str(name or "").casefold()
    if "экстра брют" in lowered:
        return "Экстра брют"
    if "брют" in lowered:
        return "Брют"
    if "полусух" in lowered:
        return "Полусухое"
    if "сух" in lowered:
        return "Сухое"
    if "полуслад" in lowered:
        return "Полусладкое"
    if "слад" in lowered:
        return "Сладкое"
    return ""


def _extract_color_and_years(name: str) -> str:
    lowered = str(name or "").casefold()
    parts: list[str] = []
    if "бел" in lowered:
        parts.append("Белое")
    elif "крас" in lowered:
        parts.append("Красное")
    elif "роз" in lowered:
        parts.append("Розовое")
    years = re.findall(r"\b(?:19|20)\d{2}\b", str(name or ""))
    parts.extend(years)
    return " ".join(parts)


def _extract_volume_liters(product: Product) -> float | str:
    match = re.search(r"(\d+(?:[.,]\d+)?)\s*(мл|ml|л|l)\b", product.name, flags=re.IGNORECASE)
    if not match:
        return ""
    value = float(match.group(1).replace(",", "."))
    if match.group(2).casefold() in {"мл", "ml"}:
        return round(value / 1000, 3)
    return value
