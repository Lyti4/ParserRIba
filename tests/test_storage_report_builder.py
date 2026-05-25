from pathlib import Path

from openpyxl import load_workbook

from models.report_request import ExportSelection, ProductFilter, ReportRequest
from models.schemas import Product
from utils.product_storage import ProductStorage
from utils.storage_report_builder import (
    build_excel_report_from_storage,
    build_report_filter_options,
)


def test_build_excel_report_from_storage_filters_by_supplier(tmp_path: Path) -> None:
    storage = ProductStorage(tmp_path / "products.db")
    storage.save_products(
        "pyaterochka",
        [
            Product(
                id="wine-1",
                name="Р’РёРЅРѕ Free Feather Chardonnay Р±РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ РїРѕР»СѓСЃР»Р°РґРєРѕРµ Р±РµР»РѕРµ 750РјР»",
                brand="Free Feather",
                price=699.99,
                image_url="https://img.example/wine-1.webp",
                product_link="https://5ka.ru/product/wine--wine-1/",
                category="Р‘РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ РІРёРЅРѕ",
                subcategory="РўРёС…РѕРµ",
                in_stock=True,
                raw_data={
                    "supplier": "Free Feather",
                    "alcohol_type": "Р‘РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ",
                },
            ),
            Product(
                id="wine-2",
                name="Р’РёРЅРѕ OddBird Spumante Р±РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ Р±РµР»РѕРµ 750РјР»",
                brand="OddBird",
                price=899.99,
                image_url="https://img.example/wine-2.webp",
                product_link="https://5ka.ru/product/wine--wine-2/",
                category="Р‘РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ РІРёРЅРѕ",
                subcategory="РРіСЂРёСЃС‚РѕРµ",
                in_stock=True,
                raw_data={
                    "supplier": "OddBird",
                    "alcohol_type": "Р‘РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ",
                },
            ),
        ],
    )
    request = ReportRequest(
        selection=ExportSelection(
            shop="pyaterochka",
            intent="wine_catalog",
            categories=["Р‘РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ РІРёРЅРѕ"],
        ),
        filters=ProductFilter(suppliers=["Free Feather"]),
        output_name="wine_supplier_report",
    )

    result = build_excel_report_from_storage(
        request,
        db_path=tmp_path / "products.db",
        output_dir=tmp_path,
    )

    assert result.products_count == 1
    assert result.categories == ["Р‘РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ РІРёРЅРѕ"]
    assert result.report_path.exists()

    workbook = load_workbook(result.report_path, read_only=True)
    assert workbook.sheetnames == ["Сводка", "Р‘РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ РІ"]
    sheet = workbook["Р‘РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ РІ"]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[1][3] == "Р’РёРЅРѕ Free Feather Chardonnay Р±РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ РїРѕР»СѓСЃР»Р°РґРєРѕРµ Р±РµР»РѕРµ 750РјР»"


def test_build_excel_report_from_storage_splits_fish_and_wine_requests(tmp_path: Path) -> None:
    storage = ProductStorage(tmp_path / "products.db")
    storage.save_products(
        "pyaterochka",
        [
            Product(
                id="fish-1",
                name="РўСЂРµСЃРєР° Р°С‚Р»Р°РЅС‚РёС‡РµСЃРєР°СЏ СЃС‚РµР№Рє Р·Р°РјРѕСЂРѕР¶РµРЅРЅС‹Р№ 600Рі",
                price=999.99,
                image_url="https://img.example/fish-1.webp",
                product_link="https://5ka.ru/product/fish--fish-1/",
                category="Р С‹Р±Р°",
                in_stock=True,
            ),
            Product(
                id="wine-1",
                name="Р’РёРЅРѕ Free Feather Chardonnay Р±РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ РїРѕР»СѓСЃР»Р°РґРєРѕРµ Р±РµР»РѕРµ 750РјР»",
                brand="Free Feather",
                price=699.99,
                image_url="https://img.example/wine-1.webp",
                product_link="https://5ka.ru/product/wine--wine-1/",
                category="Р‘РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ РІРёРЅРѕ",
                subcategory="РўРёС…РѕРµ",
                in_stock=True,
                raw_data={"alcohol_type": "Р‘РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ"},
            ),
        ],
    )

    fish_result = build_excel_report_from_storage(
        ReportRequest(
            selection=ExportSelection(
                shop="pyaterochka",
                intent="fish_catalog",
                categories=["Р С‹Р±Р°"],
            ),
            output_name="fish_report",
        ),
        db_path=tmp_path / "products.db",
        output_dir=tmp_path,
    )
    wine_result = build_excel_report_from_storage(
        ReportRequest(
            selection=ExportSelection(
                shop="pyaterochka",
                intent="wine_catalog",
                categories=["Р‘РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ РІРёРЅРѕ"],
            ),
            output_name="wine_report",
        ),
        db_path=tmp_path / "products.db",
        output_dir=tmp_path,
    )

    assert fish_result.products_count == 1
    assert wine_result.products_count == 1
    assert fish_result.report_path.name == "fish_report.xlsx"
    assert wine_result.report_path.name == "wine_report.xlsx"


def test_build_excel_report_from_storage_preserves_cyrillic_output_name(tmp_path: Path) -> None:
    storage = ProductStorage(tmp_path / "products.db")
    storage.save_products(
        "pyaterochka",
        [
            Product(
                id="fish-1",
                name="РўСЂРµСЃРєР° Р°С‚Р»Р°РЅС‚РёС‡РµСЃРєР°СЏ СЃС‚РµР№Рє Р·Р°РјРѕСЂРѕР¶РµРЅРЅС‹Р№ 600Рі",
                price=999.99,
                image_url="https://img.example/fish-1.webp",
                product_link="https://5ka.ru/product/fish--fish-1/",
                category="Р С‹Р±Р°",
                in_stock=True,
            ),
        ],
    )

    result = build_excel_report_from_storage(
        ReportRequest(
            selection=ExportSelection(
                shop="pyaterochka",
                intent="fish_catalog",
                categories=["Р С‹Р±Р°"],
            ),
            output_name="РѕС‚С‡РµС‚_СЂС‹Р±Р°",
        ),
        db_path=tmp_path / "products.db",
        output_dir=tmp_path,
    )

    assert result.report_path.name == "РѕС‚С‡РµС‚_СЂС‹Р±Р°.xlsx"


def test_build_report_filter_options_collects_supplier_and_wine_facets(tmp_path: Path) -> None:
    storage = ProductStorage(tmp_path / "products.db")
    storage.save_products(
        "pyaterochka",
        [
            Product(
                id="wine-1",
                name="Р’РёРЅРѕ Free Feather Chardonnay Р±РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ РїРѕР»СѓСЃР»Р°РґРєРѕРµ Р±РµР»РѕРµ 750РјР»",
                brand="Free Feather",
                price=699.99,
                image_url="https://img.example/wine-1.webp",
                product_link="https://5ka.ru/product/wine--wine-1/",
                category="Р‘РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ РІРёРЅРѕ",
                subcategory="РўРёС…РѕРµ",
                in_stock=True,
                raw_data={
                    "supplier": "Free Feather",
                    "alcohol_type": "Р‘РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ",
                },
            ),
            Product(
                id="wine-2",
                name="Р’РёРЅРѕ OddBird Spumante Р±РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ Р±РµР»РѕРµ 750РјР»",
                brand="OddBird",
                price=899.99,
                image_url="https://img.example/wine-2.webp",
                product_link="https://5ka.ru/product/wine--wine-2/",
                category="Р‘РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ РІРёРЅРѕ",
                subcategory="РРіСЂРёСЃС‚РѕРµ",
                in_stock=True,
                raw_data={
                    "supplier": "OddBird",
                    "alcohol_type": "Р‘РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ",
                },
            ),
        ],
    )

    result = build_report_filter_options(
        ReportRequest(
            selection=ExportSelection(
                shop="pyaterochka",
                intent="wine_catalog",
                categories=["Р‘РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ РІРёРЅРѕ"],
            )
        ),
        db_path=tmp_path / "products.db",
    )

    assert result.products_count == 2
    assert result.categories == ["Р‘РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ РІРёРЅРѕ"]
    assert result.available_filters["suppliers"] == ["Free Feather", "OddBird"]
    assert result.available_filters["brands"] == ["Free Feather", "OddBird"]
    assert result.available_filters["alcohol_types"] == ["Р‘РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ"]
    assert result.available_filters["colors"] == []
    assert result.available_filter_counts["suppliers"] == {"Free Feather": 1, "OddBird": 1}
    assert result.available_filter_counts["brands"] == {"Free Feather": 1, "OddBird": 1}
    assert result.available_filter_counts["alcohol_types"] == {"Р‘РµР·Р°Р»РєРѕРіРѕР»СЊРЅРѕРµ": 2}
    assert result.available_filter_counts["colors"] == {}


def test_build_excel_report_from_storage_prefers_selected_product_ids(tmp_path: Path) -> None:
    storage = ProductStorage(tmp_path / "products.db")
    storage.save_products(
        "pyaterochka",
        [
            Product(
                id="fish-1",
                name="Треска",
                price=199.99,
                image_url="https://img.example/fish-1.webp",
                product_link="https://5ka.ru/product/fish-1/",
                category="Рыба",
                in_stock=True,
            ),
            Product(
                id="fish-2",
                name="Лосось",
                price=299.99,
                image_url="https://img.example/fish-2.webp",
                product_link="https://5ka.ru/product/fish-2/",
                category="Рыба",
                in_stock=True,
            ),
        ],
    )

    result = build_excel_report_from_storage(
        ReportRequest(
            selection=ExportSelection(
                shop="pyaterochka",
                intent="fish_catalog",
                categories=["Рыба"],
                selected_product_ids=["fish-2"],
            ),
            output_name="selected_product_report",
        ),
        db_path=tmp_path / "products.db",
        output_dir=tmp_path,
    )

    assert result.products_count == 1
    workbook = load_workbook(result.report_path, read_only=True)
    sheet = workbook["Рыба"]
    rows = list(sheet.iter_rows(values_only=True))
    assert rows[1][3] == "Лосось"
